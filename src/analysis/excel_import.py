"""
Import LLM-filled Excel analysis results back into the database.

Supports results produced by any provider (OpenAI, Claude, Gemini, etc.)
via the ChatGPT or Claude.ai web interface — or the API.

Usage:
    from src.analysis.excel_import import import_results
    stats = import_results(excel_path, "classify", "openai", "gpt-5.4-thinking")
"""
import json
from pathlib import Path
from typing import Optional

try:
    import openpyxl
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

from src.analysis.runner import (
    create_run,
    finish_run,
    store_result,
    store_topics_for_email,
    store_timeline_events,
)

# ── Provider name normalisation ───────────────────────────────────────────────

_PROVIDER_ALIASES: dict = {
    "chatgpt":   "openai",
    "gpt":       "openai",
    "openai":    "openai",
    "claude":    "claude",
    "anthropic": "claude",
    "gemini":    "google",
    "google":    "google",
    "mistral":   "mistral",
    "llama":     "meta",
    "meta":      "meta",
}

_MODEL_DEFAULTS: dict = {
    "openai":  "gpt-4o",
    "claude":  "claude-opus-4-5",
    "google":  "gemini-1.5-pro",
    "mistral": "mistral-large",
    "meta":    "llama-3.3-70b",
}


def _normalise_provider(provider: str) -> str:
    return _PROVIDER_ALIASES.get(provider.lower(), provider.lower())


def _default_model(provider: str) -> str:
    return _MODEL_DEFAULTS.get(provider, provider)


# ── Main entry point ──────────────────────────────────────────────────────────

def import_results(
    excel_path: Path,
    analysis_type: str,
    provider: str,
    model: Optional[str] = None,
) -> dict:
    """
    Read a filled Excel export file and insert results into the database.

    Creates a new analysis_run tagged with the given provider/model.
    Idempotent: rows with empty output columns are silently skipped.

    Returns a stats dict:
        {run_id, imported, skipped, errors, status}
    """
    if not _HAS_OPENPYXL:
        raise ImportError("openpyxl required: pip install openpyxl")

    provider = _normalise_provider(provider)
    model = model or _default_model(provider)

    wb = openpyxl.load_workbook(str(excel_path), data_only=True)

    if "Emails" not in wb.sheetnames:
        raise ValueError("Workbook has no 'Emails' sheet — was this exported by MyYahooEmails?")

    # Read metadata from _meta sheet if present
    _meta = {}
    if "_meta" in wb.sheetnames:
        ws_meta = wb["_meta"]
        for row in ws_meta.iter_rows(min_row=1, values_only=True):
            if row[0] and row[1]:
                _meta[row[0]] = row[1]

    stored_type = _meta.get("analysis_type", analysis_type)
    if stored_type != analysis_type:
        raise ValueError(
            f"File was exported as '{stored_type}' but you passed --type {analysis_type}. "
            "Re-run with the correct --type."
        )

    ws = wb["Emails"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # Create a new run record tagged with this provider / model
    prompt_note = f"excel-import from {excel_path.name}"
    run_id = create_run(
        analysis_type=analysis_type,
        provider_name=provider,
        model_id=model,
        prompt_text=prompt_note,
        prompt_version="v1-excel",
        notes=prompt_note,
    )

    imported = 0
    skipped  = 0
    errors   = []

    # Contradictions and legal_analysis use separate sheet/storage paths
    if analysis_type == "contradictions":
        return _import_contradictions(wb, excel_path, run_id, provider, model)

    if analysis_type == "legal_analysis":
        return _import_legal_analysis(wb, excel_path, run_id)

    parsers = {
        "classify":     _parse_classify,
        "tone":         _parse_tone,
        "timeline":     _parse_timeline,
        "manipulation": _parse_manipulation,
    }
    parser = parsers.get(analysis_type)
    if parser is None:
        raise ValueError(f"Unsupported analysis_type: {analysis_type}")

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        try:
            email_id = int(row[0])
        except (TypeError, ValueError):
            continue

        try:
            result = parser(row, headers)

            if result is None:
                skipped += 1
                continue

            # Store raw JSON result (same schema as LLM-generated results)
            store_result(run_id, email_id, json.dumps(result))

            # Post-process into typed tables
            if analysis_type == "classify" and result.get("topics"):
                store_topics_for_email(email_id, result["topics"], run_id)

            elif analysis_type == "timeline" and result.get("events"):
                store_timeline_events(run_id, email_id, result["events"])

            imported += 1

        except Exception as exc:
            errors.append(f"email_id={email_id}: {exc}")

    status = "complete" if not errors else "partial"
    finish_run(run_id, status, imported)

    return {
        "run_id":   run_id,
        "imported": imported,
        "skipped":  skipped,
        "errors":   errors,
        "status":   status,
    }


# ── Row parsers ───────────────────────────────────────────────────────────────

def _get(row, headers, name):
    """Return the value for a given column name, or None."""
    try:
        idx = headers.index(name)
        val = row[idx] if idx < len(row) else None
        return str(val).strip() if val is not None else None
    except ValueError:
        return None


def _parse_classify(row, headers) -> Optional[dict]:
    topics_raw = _get(row, headers, "topics")
    if not topics_raw:
        return None

    confidence_raw = _get(row, headers, "confidence") or ""
    summary        = _get(row, headers, "summary") or ""

    topic_names = [t.strip() for t in topics_raw.split(",") if t.strip()]
    conf_values = [c.strip() for c in confidence_raw.split(",") if c.strip()]

    topics = []
    for i, name in enumerate(topic_names):
        try:
            conf = float(conf_values[i]) if i < len(conf_values) else 0.8
            conf = max(0.0, min(1.0, conf))
        except (ValueError, IndexError):
            conf = 0.8
        topics.append({"name": name, "confidence": round(conf, 2)})

    return {
        "topics":  topics,
        "summary": summary,
        "source":  "excel-import",
    }


def _parse_tone(row, headers) -> Optional[dict]:
    tone = _get(row, headers, "tone")
    if not tone:
        return None

    def to_float(name):
        val = _get(row, headers, name)
        try:
            return round(max(0.0, min(1.0, float(val))), 2)
        except (TypeError, ValueError):
            return 0.0

    return {
        "tone":               tone,
        "aggression_level":   to_float("aggression_level"),
        "manipulation_score": to_float("manipulation_score"),
        "legal_posturing":    to_float("legal_posturing"),
        "summary":            _get(row, headers, "summary") or "",
        "source":             "excel-import",
    }


def _parse_timeline(row, headers) -> Optional[dict]:
    description = _get(row, headers, "description")

    # Blank = reviewed, no extractable event found — store as empty events list
    if not description:
        return {"events": [], "source": "excel-import"}

    event_date = _get(row, headers, "event_date") or ""
    # Normalise date: keep first 10 chars (YYYY-MM-DD) if present.
    # Pad partial dates so SQLite TIMESTAMP affinity doesn't coerce "2015" → int.
    if event_date and len(event_date) >= 10:
        event_date = event_date[:10]
    elif event_date and len(event_date) == 4 and event_date.isdigit():
        event_date = event_date + "-01"   # "2015" → "2015-01"
    elif event_date and len(event_date) == 7 and event_date[4] == "-":
        pass  # "2015-01" already safe — no coercion possible

    return {
        "events": [{
            "event_date":  event_date,
            "event_type":  _get(row, headers, "event_type") or "statement",
            "significance": _get(row, headers, "significance") or "medium",
            "description":  description,
        }],
        "source": "excel-import",
    }


def _parse_manipulation(row, headers) -> Optional[dict]:
    """Parse a manipulation row: total_score, dominant_pattern, detected_patterns, notes.

    Blank rows are stored as total_score=0.0 (clean — reviewed, no manipulation detected).
    This ensures every processed email is tracked and won't reappear in future exports.
    """
    total_score_raw = _get(row, headers, "total_score")

    # Blank = intentionally left blank by LLM = no manipulation detected = score 0.0
    if not total_score_raw:
        return {
            "patterns":         [],
            "total_score":      0.0,
            "dominant_pattern": None,
            "notes":            "",
            "source":           "excel-import",
        }

    try:
        total_score = round(max(0.0, min(1.0, float(total_score_raw))), 2)
    except (TypeError, ValueError):
        return None

    dominant_pattern = _get(row, headers, "dominant_pattern") or None
    detected_raw     = _get(row, headers, "detected_patterns") or ""
    notes            = _get(row, headers, "notes") or ""

    # Parse "gaslighting:0.8, projection:0.5" → [{type, score}]
    patterns = []
    for item in detected_raw.split(","):
        item = item.strip()
        if not item:
            continue
        if ":" in item:
            parts = item.split(":", 1)
            pattern_type = parts[0].strip()
            try:
                score = round(max(0.0, min(1.0, float(parts[1].strip()))), 2)
            except (ValueError, IndexError):
                score = 0.5
        else:
            pattern_type = item
            score = 0.5
        if pattern_type:
            patterns.append({"type": pattern_type, "score": score,
                             "evidence": "", "explanation": ""})

    return {
        "patterns":        patterns,
        "total_score":     total_score,
        "dominant_pattern": dominant_pattern,
        "notes":           notes,
        "source":          "excel-import",
    }


def _import_legal_analysis(wb, excel_path: Path, run_id: int) -> dict:
    """
    Import legal_analysis results from a filled workbook.

    Events sheet  → procedure_events (procedure_ref is a valid int 1–N)
                    or timeline_events (no linked procedure)
    Analysis sheet → analysis_results as JSON, one row per email.
    Uses a single connection for all writes (no nested get_db() calls).
    """
    from src.storage.database import get_db

    if "Events" not in wb.sheetnames:
        raise ValueError("No 'Events' sheet found. Was this exported with --type legal_analysis?")
    if "Analysis" not in wb.sheetnames:
        raise ValueError("No 'Analysis' sheet found. Was this exported with --type legal_analysis?")

    events_imported   = 0
    analysis_imported = 0
    skipped           = 0
    errors            = []

    def _cell(row, headers, name):
        """Return stripped string value for column `name`, or None if blank."""
        try:
            idx = headers.index(name)
            val = row[idx] if idx < len(row) else None
            return str(val).strip() if val not in (None, "") else None
        except ValueError:
            return None

    def _norm_date(raw):
        """Normalise a date string to YYYY-MM-DD (or shorter, as-is)."""
        if not raw:
            return None
        raw = str(raw).strip()
        if len(raw) >= 10:
            return raw[:10]
        if len(raw) == 4 and raw.isdigit():
            return raw + "-01-01"
        if len(raw) == 7 and raw[4] == "-":
            return raw + "-01"
        return raw or None

    with get_db() as conn:
        valid_proc_ids = {
            r[0] for r in conn.execute("SELECT id FROM procedures").fetchall()
        }

        # ── Email date cache (fallback when LLM leaves event_date blank) ────────
        email_date_cache = {}

        def _email_date(eid):
            if eid not in email_date_cache:
                row_d = conn.execute(
                    "SELECT date FROM emails WHERE id=?", (eid,)
                ).fetchone()
                email_date_cache[eid] = str(row_d[0])[:10] if row_d and row_d[0] else "1900-01-01"
            return email_date_cache[eid]

        # ── Events sheet ──────────────────────────────────────────────────────
        ws_ev = wb["Events"]
        ev_hdrs = [cell.value for cell in next(ws_ev.iter_rows(min_row=1, max_row=1))]

        for row in ws_ev.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            try:
                email_id = int(row[0])
            except (TypeError, ValueError):
                continue

            description = _cell(row, ev_hdrs, "description")
            if not description:
                skipped += 1
                continue

            event_type   = _cell(row, ev_hdrs, "event_type") or "legal"
            event_date   = _norm_date(_cell(row, ev_hdrs, "event_date"))
            proc_raw     = _cell(row, ev_hdrs, "procedure_ref")
            amount_raw   = _cell(row, ev_hdrs, "amount_eur")
            significance = _cell(row, ev_hdrs, "significance") or "medium"

            # When LLM omits the date, fall back to the email's own date
            # and mark precision as "approximate" so it's distinguishable.
            if not event_date:
                event_date = _email_date(email_id)
                date_precision = "approximate"
            else:
                date_precision = "exact" if len(event_date) == 10 else "approximate"

            try:
                proc_id = int(proc_raw) if proc_raw and str(proc_raw).isdigit() else None
                if proc_id and proc_id not in valid_proc_ids:
                    proc_id = None

                if proc_id:
                    conn.execute(
                        """INSERT INTO procedure_events
                           (procedure_id, event_date, event_type, date_precision,
                            description, outcome, jurisdiction, source_email_id, notes)
                           VALUES (?, ?, ?, ?, ?, '', '', ?, ?)""",
                        (
                            proc_id, event_date, event_type, date_precision,
                            description, email_id,
                            f"excel-import run_id={run_id} significance={significance}"
                            + (f" amount={amount_raw}EUR" if amount_raw else ""),
                        ),
                    )
                else:
                    conn.execute(
                        """INSERT INTO timeline_events
                           (run_id, email_id, event_date, event_type, description, significance)
                           VALUES (?, ?, ?, ?, ?, ?)""",
                        (run_id, email_id, event_date, event_type, description, significance),
                    )
                events_imported += 1

            except Exception as exc:
                errors.append(f"Events email_id={email_id}: {exc}")

        # ── Analysis sheet ────────────────────────────────────────────────────
        ws_an = wb["Analysis"]
        an_hdrs = [cell.value for cell in next(ws_an.iter_rows(min_row=1, max_row=1))]

        for row in ws_an.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            try:
                email_id = int(row[0])
            except (TypeError, ValueError):
                continue

            filled = [row[i] for i in range(1, min(len(row), len(an_hdrs)))]
            if not any(v not in (None, "", " ") for v in filled):
                skipped += 1
                continue

            try:
                mood_intensity_raw = _cell(row, an_hdrs, "mood_intensity")
                try:
                    mood_intensity = int(float(mood_intensity_raw)) if mood_intensity_raw else None
                except (ValueError, TypeError):
                    mood_intensity = None

                result = {
                    "mood_valence":       _cell(row, an_hdrs, "mood_valence"),
                    "mood_intensity":     mood_intensity,
                    "urgency":            _cell(row, an_hdrs, "urgency"),
                    "key_concern":        _cell(row, an_hdrs, "key_concern"),
                    "trust_in_lawyer":    _cell(row, an_hdrs, "trust_in_lawyer"),
                    "father_role_stress": _cell(row, an_hdrs, "father_role_stress"),
                    "financial_stress":   _cell(row, an_hdrs, "financial_stress"),
                    "lawyer_stance":      _cell(row, an_hdrs, "lawyer_stance"),
                    "strategy_signal":    _cell(row, an_hdrs, "strategy_signal"),
                    "action_required":    _cell(row, an_hdrs, "action_required"),
                    "risk_signal":        _cell(row, an_hdrs, "risk_signal"),
                    "procedure_ref":      _cell(row, an_hdrs, "procedure_ref"),
                    "persons_mentioned":  _cell(row, an_hdrs, "persons_mentioned"),
                    "amounts_mentioned":  _cell(row, an_hdrs, "amounts_mentioned"),
                    "children_mentioned": _cell(row, an_hdrs, "children_mentioned"),
                    "source":             "excel-import",
                }
                result = {k: v for k, v in result.items() if v is not None}

                conn.execute(
                    """INSERT OR REPLACE INTO analysis_results
                       (run_id, email_id, sender_contact_id, result_json)
                       VALUES (?, ?, NULL, ?)""",
                    (run_id, email_id, json.dumps(result)),
                )
                analysis_imported += 1

            except Exception as exc:
                errors.append(f"Analysis email_id={email_id}: {exc}")

        conn.commit()

    imported = events_imported + analysis_imported
    status = "complete" if not errors else "partial"
    finish_run(run_id, status, imported)

    return {
        "run_id":            run_id,
        "events_imported":   events_imported,
        "analysis_imported": analysis_imported,
        "imported":          imported,
        "skipped":           skipped,
        "errors":            errors,
        "status":            status,
    }


def _import_contradictions(wb, excel_path: Path, run_id: int,
                            provider: str, model: str) -> dict:
    """Import contradiction pairs from the 'Contradictions' output sheet."""
    from src.storage.database import get_db

    if "Contradictions" not in wb.sheetnames:
        raise ValueError(
            "No 'Contradictions' sheet found. "
            "Make sure you filled in the Contradictions output sheet."
        )

    ws = wb["Contradictions"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    imported = 0
    skipped  = 0
    errors   = []

    with get_db() as conn:
        for row in ws.iter_rows(min_row=3, values_only=True):  # skip header + hint row
            if not row or not row[0]:
                skipped += 1
                continue

            def g(name):
                try:
                    idx = headers.index(name)
                    val = row[idx] if idx < len(row) else None
                    return str(val).strip() if val is not None else None
                except ValueError:
                    return None

            email_id_a  = g("email_id_a")
            email_id_b  = g("email_id_b")
            explanation = g("explanation")

            if not email_id_a or not email_id_b or not explanation:
                skipped += 1
                continue

            # Skip hint row values
            if email_id_a.startswith("←") or not email_id_a.isdigit():
                skipped += 1
                continue

            try:
                conn.execute(
                    """INSERT OR IGNORE INTO contradictions
                       (run_id, email_id_a, email_id_b, scope, topic, severity, explanation)
                       VALUES (?, ?, ?, ?, ?, ?, ?)""",
                    (
                        run_id,
                        int(email_id_a),
                        int(email_id_b),
                        g("scope") or "intra-sender",
                        g("topic") or None,
                        g("severity") or "medium",
                        explanation,
                    ),
                )
                # Also store a result_json for traceability on email_id_a
                result = json.dumps({
                    "email_id_a": int(email_id_a),
                    "email_id_b": int(email_id_b),
                    "scope":      g("scope"),
                    "topic":      g("topic"),
                    "severity":   g("severity"),
                    "explanation": explanation,
                    "source":     "excel-import",
                })
                conn.execute(
                    """INSERT OR REPLACE INTO analysis_results
                       (run_id, email_id, sender_contact_id, result_json)
                       VALUES (?, ?, NULL, ?)""",
                    (run_id, int(email_id_a), result),
                )
                imported += 1
            except Exception as exc:
                errors.append(f"row ({email_id_a},{email_id_b}): {exc}")

        conn.commit()

    status = "complete" if not errors else "partial"
    finish_run(run_id, status, imported)

    return {
        "run_id":   run_id,
        "imported": imported,
        "skipped":  skipped,
        "errors":   errors,
        "status":   status,
    }
