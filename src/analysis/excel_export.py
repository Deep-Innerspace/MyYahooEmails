"""
Export emails to Excel for manual LLM analysis (ChatGPT / Claude web interface).

Usage:
    from src.analysis.excel_export import export_for_analysis
    path, count = export_for_analysis(conn, "classify", Path("exports/batch.xlsx"), limit=10)
"""
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

# ── Analysis type definitions ─────────────────────────────────────────────────

ANALYSIS_TYPES = {
    "classify": {
        "title": "Topic Classification",
        "task": (
            "Classify each French email by topic. "
            "Fill the YELLOW columns for every row. "
            "Use only topics from the 'Available Topics' list below. "
            "Multiple topics allowed per email, comma-separated. "
            "Do not modify any blue (input) columns."
        ),
        "output_columns": [
            ("topics",     "Comma-separated topic names from the allowed list  e.g.  enfants, finances"),
            ("confidence", "Comma-separated confidence scores 0.0–1.0 matching topics order  e.g.  0.9, 0.75"),
            ("summary",    "1–2 sentence summary of the email content in French"),
        ],
    },
    "tone": {
        "title": "Tone Analysis",
        "task": (
            "Analyse the tone, aggression, and manipulation level of each French email. "
            "Fill the YELLOW columns for every row. "
            "Scores are 0.0 (none) to 1.0 (extreme). "
            "Do not modify any blue (input) columns."
        ),
        "output_columns": [
            ("tone",              "Overall tone label: neutre, cordial, agressif, manipulateur, victimisant, menaçant"),
            ("aggression_level",  "Aggression score 0.0–1.0"),
            ("manipulation_score","Manipulation score 0.0–1.0"),
            ("legal_posturing",   "Legal posturing score 0.0–1.0  (threats, formal citations, pressure tactics)"),
            ("summary",           "1–2 sentence analysis note in French"),
        ],
    },
    "timeline": {
        "title": "Timeline Event Extraction",
        "task": (
            "Extract the key dated event from each French email (one event maximum per row). "
            "Leave all output columns blank if the email contains no extractable event. "
            "Do not modify any blue (input) columns."
        ),
        "output_columns": [
            ("event_date",  "Date of the event mentioned: YYYY-MM-DD format, or leave blank"),
            ("event_type",  "Type: statement | demand | agreement | threat | legal | financial | personal"),
            ("significance","Significance: high | medium | low"),
            ("description", "Event description in French (1 sentence)"),
        ],
    },
    "manipulation": {
        "title": "Manipulation Pattern Detection",
        "task": (
            "Detect manipulation tactics in each French email (divorce legal context). "
            "Fill the YELLOW columns for every row. Score 0.0=absent to 1.0=extreme. "
            "Leave all 4 columns blank if no manipulation is detected (total_score=0). "
            "Do not modify any blue (input) columns."
        ),
        "output_columns": [
            ("total_score",       "Overall manipulation score 0.0–1.0  (0.0 if none detected)"),
            ("dominant_pattern",  "Main pattern detected — one of: gaslighting | emotional_weaponization | "
                                  "financial_coercion | legal_threats | children_instrumentalization | "
                                  "guilt_tripping | projection | false_victimhood | moving_goalposts | "
                                  "silent_treatment_threat — or leave blank"),
            ("detected_patterns", "Comma-separated list of detected patterns with scores: "
                                  "e.g.  gaslighting:0.8, projection:0.5  — leave blank if none"),
            ("notes",             "1–2 sentence evaluation in French (or leave blank)"),
        ],
    },
}

# ── Colour palette ────────────────────────────────────────────────────────────

_NAVY      = "1E3A5F"
_ORANGE    = "C0600A"
_YELLOW_BG = "FFF9C4"
_BLUE_HDR  = "2C5282"
_WHITE     = "FFFFFF"
_GRAY      = "888888"


_EXCEL_CELL_LIMIT = 32767


def export_for_analysis(
    conn: sqlite3.Connection,
    analysis_type: str,
    output_path: Path,
    limit: Optional[int] = None,
    offset: int = 0,
    unanalyzed_only: bool = True,
    exclude_large: bool = True,
) -> tuple:
    """
    Export emails to an Excel workbook for manual LLM analysis.

    Returns:
        (output_path, email_count)
    """
    if not _HAS_OPENPYXL:
        raise ImportError("openpyxl required: pip install openpyxl")

    if analysis_type not in ANALYSIS_TYPES:
        raise ValueError(
            f"Unknown analysis type '{analysis_type}'. "
            f"Choose from: {', '.join(ANALYSIS_TYPES)}"
        )

    cfg = ANALYSIS_TYPES[analysis_type]
    output_cols = cfg["output_columns"]

    # ── Fetch emails ──────────────────────────────────────────────────────────
    wheres = [
        "e.corpus = 'personal'",
        "e.delta_text IS NOT NULL",
        "LENGTH(TRIM(e.delta_text)) > 10",
    ]

    if exclude_large:
        wheres.append(f"LENGTH(e.delta_text) <= {_EXCEL_CELL_LIMIT}")

    if unanalyzed_only:
        if analysis_type == "classify":
            wheres.append(
                "e.id NOT IN (SELECT DISTINCT email_id FROM email_topics)"
            )
        else:
            wheres.append(f"""e.id NOT IN (
                SELECT DISTINCT ar.email_id
                FROM analysis_results ar
                JOIN analysis_runs ru ON ru.id = ar.run_id
                WHERE ru.analysis_type = '{analysis_type}'
                  AND ru.status IN ('complete', 'partial')
            )""")

    if limit:
        limit_sql = f"LIMIT {limit} OFFSET {offset}"
    elif offset:
        limit_sql = f"LIMIT -1 OFFSET {offset}"
    else:
        limit_sql = ""

    rows = conn.execute(f"""
        SELECT e.id, e.date, e.direction, e.subject, e.delta_text,
               c.name AS contact_name
        FROM emails e
        LEFT JOIN contacts c ON c.id = e.contact_id
        WHERE {' AND '.join(wheres)}
        ORDER BY e.date ASC
        {limit_sql}
    """).fetchall()

    if not rows:
        raise ValueError(
            "No emails to export — all are already analyzed, "
            "or no delta_text available."
        )

    # ── Available topics (classify only) ─────────────────────────────────────
    topics_list = []
    if analysis_type == "classify":
        topics_list = [
            r["name"]
            for r in conn.execute("SELECT name FROM topics ORDER BY name").fetchall()
        ]

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    _build_instructions_sheet(wb, cfg, output_cols, topics_list, len(rows), analysis_type)
    truncated_count = _build_emails_sheet(wb, rows, output_cols)
    _build_meta_sheet(wb, analysis_type, len(rows), output_cols)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path, len(rows), truncated_count


# ── Sheet builders ────────────────────────────────────────────────────────────

def _build_instructions_sheet(wb, cfg, output_cols, topics_list, email_count, analysis_type):
    ws = wb.active
    ws.title = "Instructions"

    def h(row, col, value, bold=False, size=11, color="000000", bg=None, wrap=False):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = Font(bold=bold, size=size, color=color)
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        if wrap:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        return cell

    # Title
    h(1, 1, f"MyYahooEmails — {cfg['title']} Batch",
      bold=True, size=15, color=_NAVY)
    h(2, 1,
      f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}   |   "
      f"{email_count} emails   |   "
      f"Import: python cli.py analyze import-results <file.xlsx> --type {analysis_type}",
      color=_GRAY)

    # Task
    h(4, 1, "TASK", bold=True, color=_NAVY)
    cell = ws.cell(row=4, column=2, value=cfg["task"])
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.font = Font(size=11)
    ws.row_dimensions[4].height = 55

    # Output columns
    h(6, 1, "COLUMNS TO FILL (yellow)", bold=True, color=_ORANGE)
    for i, (col_name, description) in enumerate(output_cols, start=7):
        h(i, 1, col_name, bold=True)
        h(i, 2, description, wrap=True)
        ws.row_dimensions[i].height = 20

    # Available topics
    next_row = 7 + len(output_cols) + 1
    if topics_list:
        h(next_row, 1, "AVAILABLE TOPICS", bold=True, color=_NAVY)
        h(next_row, 2, ", ".join(topics_list), bold=True)
        h(next_row + 1, 2,
          "Use ONLY topics from this list. "
          "New topics will be auto-created on import if not found.")
        next_row += 3

    # Provider note
    h(next_row, 1, "PROVIDER / MODEL", bold=True, color=_NAVY)
    h(next_row, 2,
      "When importing, specify --provider and --model. "
      "Examples:  --provider openai --model gpt-5.4-thinking   "
      "or  --provider claude --model claude-opus-4-5")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 90


def _build_emails_sheet(wb, rows, output_cols):
    ws = wb.create_sheet("Emails")

    input_cols = ["email_id", "date", "direction", "contact", "subject", "delta_text"]
    out_col_names = [c[0] for c in output_cols]
    all_cols = input_cols + out_col_names

    # Header row
    for col_idx, col_name in enumerate(all_cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        is_output = col_name in out_col_names
        cell.fill = PatternFill("solid", fgColor=_ORANGE if is_output else _BLUE_HDR)
        cell.font = Font(color=_WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Data rows
    yellow = PatternFill("solid", fgColor=_YELLOW_BG)
    _TRUNCATION_MARKER = "\n\n[... TRUNCATED — email exceeds Excel 32,767 char limit. Full text available in DB.]"
    truncated_count = 0

    for row_idx, row in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=row["id"])
        ws.cell(row=row_idx, column=2, value=str(row["date"])[:10] if row["date"] else "")
        ws.cell(row=row_idx, column=3, value=row["direction"] or "")
        ws.cell(row=row_idx, column=4, value=row["contact_name"] or "")
        ws.cell(row=row_idx, column=5, value=row["subject"] or "(no subject)")

        delta = row["delta_text"] or ""
        if len(delta) > _EXCEL_CELL_LIMIT:
            keep = _EXCEL_CELL_LIMIT - len(_TRUNCATION_MARKER)
            delta = delta[:keep] + _TRUNCATION_MARKER
            truncated_count += 1


        delta_cell = ws.cell(row=row_idx, column=6, value=delta)
        delta_cell.alignment = Alignment(wrap_text=False, vertical="top")

        for out_idx in range(len(out_col_names)):
            cell = ws.cell(row=row_idx, column=7 + out_idx, value="")
            cell.fill = yellow

    # Column widths
    widths = {"A": 9, "B": 12, "C": 11, "D": 16, "E": 38, "F": 65}
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width
    for i in range(len(out_col_names)):
        ws.column_dimensions[get_column_letter(7 + i)].width = 28

    ws.freeze_panes = "G2"
    return truncated_count


def _build_meta_sheet(wb, analysis_type, email_count, output_cols):
    ws = wb.create_sheet("_meta")
    ws["A1"], ws["B1"] = "analysis_type", analysis_type
    ws["A2"], ws["B2"] = "export_date", datetime.now().isoformat()
    ws["A3"], ws["B3"] = "email_count", email_count
    ws["A4"], ws["B4"] = "schema_version", "1"
    ws["A5"], ws["B5"] = "output_columns", ",".join(c[0] for c in output_cols)


# ── Legal analysis export (legal corpus, 3-sheet format) ─────────────────────

_LEGAL_EVENT_TYPES = [
    # Billing
    "invoice_issued", "payment_request", "payment_confirmed", "fee_estimate",
    "expense_note", "cost_warning", "retainer_requested",
    # Procedural filings
    "conclusions_filed", "requete_filed", "assignation", "appeal_filed",
    "document_communicated", "constitution_avocat", "desistement",
    "signification", "consignation", "incident_filed",
    # Court
    "hearing_scheduled", "hearing_occurred", "hearing_postponed",
    "hearing_cancelled", "judgment_rendered", "ordonnance_rendered",
    "arret_rendered", "expert_appointed", "expert_report_delivered",
    "deadline_set", "mise_en_etat",
    # Strategy
    "strategy_decision", "evidence_discussed", "settlement_offer_received",
    "settlement_offer_made", "settlement_rejected", "adverse_move",
    "judge_observation", "case_assessment", "client_instruction",
    "lawyer_recommendation",
    # Admin
    "meeting_scheduled", "meeting_occurred", "document_exchange",
    "procuration", "change_of_lawyer",
]

_LEGAL_MOOD_VALUES      = "distressed | anxious | angry | frustrated | resigned | determined | hopeful | relieved | neutral"
_LEGAL_URGENCY_VALUES   = "none | moderate | urgent | panic"
_LEGAL_TRUST_VALUES     = "high | medium | low | questioning"
_LEGAL_STANCE_VALUES    = "reassuring | cautious | tactical | concerned | optimistic | pessimistic | urgent | evasive"
_LEGAL_RISK_VALUES      = "none | low | medium | high | critical"
_LEGAL_CHILDREN_VALUES  = "none | Matheys | Lounys | Maylis | multiple"

# Content threshold: if delta_text shorter than this, fall back to body_text
_DELTA_MIN_CHARS = 150
# Truncation limit — leave headroom so the marker fits within the 32,767 Excel cell limit
_LEGAL_CELL_LIMIT = 30000


def export_legal_analysis(
    conn: sqlite3.Connection,
    output_path: Path,
    limit: Optional[int] = 150,
    offset: int = 0,
    unanalyzed_only: bool = True,
) -> tuple:
    """
    Export legal corpus emails to Excel for comprehensive ChatGPT analysis.

    Three output sheets:
      - Emails   : blue read-only input  (email_id, date, direction, contact, subject, content)
      - Events   : orange/yellow output  (0..N events per email — email_id pre-filled)
      - Analysis : orange/yellow output  (exactly 1 metadata row per email — email_id pre-filled)

    Content field: delta_text if ≥ 150 chars, else body_text (fallback for short/stripped deltas).
    Emails whose combined content exceeds 30,000 chars are truncated with a visible marker.

    Returns (output_path, email_count, truncated_count).
    """
    if not _HAS_OPENPYXL:
        raise ImportError("openpyxl required: pip install openpyxl")

    # ── Fetch procedures & lawyers for instructions ───────────────────────────
    procedures = conn.execute(
        "SELECT id, name, case_number FROM procedures ORDER BY id"
    ).fetchall()

    lawyers = conn.execute(
        "SELECT name, role FROM contacts "
        "WHERE role IN ('my_lawyer','her_lawyer','opposing_counsel','notaire') "
        "ORDER BY role, name"
    ).fetchall()

    # ── Fetch legal emails ────────────────────────────────────────────────────
    wheres = ["e.corpus = 'legal'"]

    if unanalyzed_only:
        wheres.append("""e.id NOT IN (
            SELECT DISTINCT ar.email_id
            FROM analysis_results ar
            JOIN analysis_runs ru ON ru.id = ar.run_id
            WHERE ru.analysis_type = 'legal_analysis'
              AND ru.status IN ('complete', 'partial')
        )""")

    limit_sql = ""
    if limit:
        limit_sql = f"LIMIT {limit} OFFSET {offset}"
    elif offset:
        limit_sql = f"LIMIT -1 OFFSET {offset}"

    rows = conn.execute(f"""
        SELECT e.id, e.date, e.direction, e.subject,
               e.delta_text, e.body_text,
               c.name AS contact_name, c.role AS contact_role
        FROM emails e
        LEFT JOIN contacts c ON c.id = e.contact_id
        WHERE {' AND '.join(wheres)}
        ORDER BY e.date ASC
        {limit_sql}
    """).fetchall()

    if not rows:
        raise ValueError(
            "No legal corpus emails to export — all are already analyzed, "
            "or no legal emails in the database."
        )

    # ── Build content field for each row ─────────────────────────────────────
    _TRUNC_MARKER = (
        "\n\n[... TRUNCATED — email exceeds 30,000 chars. Full text available in DB.]"
    )
    email_data = []
    truncated_count = 0
    for row in rows:
        delta   = (row["delta_text"] or "").strip()
        body    = (row["body_text"]  or "").strip()
        content = delta if len(delta) >= _DELTA_MIN_CHARS else (body or delta)
        if len(content) > _LEGAL_CELL_LIMIT:
            content = content[:_LEGAL_CELL_LIMIT] + _TRUNC_MARKER
            truncated_count += 1
        email_data.append({
            "id":           row["id"],
            "date":         str(row["date"])[:10] if row["date"] else "",
            "direction":    row["direction"] or "",
            "contact_name": row["contact_name"] or "(inconnu)",
            "contact_role": row["contact_role"] or "",
            "subject":      row["subject"] or "(no subject)",
            "content":      content,
        })

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    _build_legal_instructions_sheet(wb, len(email_data), procedures, lawyers)
    _build_legal_emails_sheet(wb, email_data)
    _build_legal_events_sheet(wb, email_data)
    _build_legal_analysis_sheet(wb, email_data)
    _build_legal_meta_sheet(wb, len(email_data))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path, len(email_data), truncated_count


# ── Legal sheet builders ──────────────────────────────────────────────────────

def _build_legal_instructions_sheet(wb, email_count, procedures, lawyers):
    ws = wb.active
    ws.title = "Instructions"

    def w(row, col, value, bold=False, size=11, color="000000", bg=None, wrap=False, italic=False):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = Font(bold=bold, size=size, color=color, italic=italic)
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        if wrap:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        return cell

    r = 1
    w(r, 1, "MyYahooEmails — Legal Corpus Analysis", bold=True, size=16, color=_NAVY)
    r += 1
    w(r, 1,
      f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
      f"{email_count} emails  |  "
      "corpus: legal  |  "
      "Import: python cli.py analyze import-results <file.xlsx> --type legal_analysis",
      color=_GRAY, size=10)

    r += 2
    w(r, 1, "CONTEXT", bold=True, color=_NAVY)
    ctx = (
        "These are emails between Gaël MAISON (architecte, Abu Dhabi) and his lawyers, "
        "opposing counsel, and a notaire — spanning a 10-year French divorce and custody "
        "battle (2014–present). The emails are predominantly in French. "
        "Gaël's lawyers: Valérie Charriot-Lecuyer (2014–2016), then Hélène Hartwig-Deblauwe "
        "and the Onyx Avocats team (2017–present), and François Teytaud for appeals. "
        "Opposing party: Maud MULLER (ex-wife). Her lawyers appear in the LAWYERS table below."
    )
    cell = ws.cell(row=r, column=2, value=ctx)
    cell.font = Font(size=11)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 55

    r += 2
    w(r, 1, "YOUR TASK", bold=True, color=_NAVY)
    task = (
        "For EACH email in the 'Emails' sheet:\n"
        "  1. Fill the 'Events' sheet — add one row per legal/financial event you detect "
        "(0 rows if no event). The email_id column is pre-filled; copy it if you add extra rows.\n"
        "  2. Fill the 'Analysis' sheet — fill the row that already has the matching email_id "
        "(exactly one row per email). Fill ALL yellow cells.\n"
        "Do NOT modify the 'Emails' sheet. Work through emails in order (top to bottom)."
    )
    cell = ws.cell(row=r, column=2, value=task)
    cell.font = Font(size=11)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 70

    r += 2
    w(r, 1, "LANGUAGE", bold=True, color=_NAVY)
    w(r, 2,
      "All description / key_concern / strategy_signal fields: write in FRENCH. "
      "All vocabulary fields (event_type, mood_valence, etc.): use the EXACT English keywords listed below.",
      wrap=True)
    ws.row_dimensions[r].height = 30

    # ── Events sheet columns ──────────────────────────────────────────────────
    r += 2
    w(r, 1, "EVENTS SHEET — COLUMNS", bold=True, color=_ORANGE)
    r += 1
    events_cols = [
        ("email_id",       "Pre-filled. Copy it for additional event rows on the same email."),
        ("event_date",     "Date of the event: YYYY-MM-DD, YYYY-MM, or YYYY. Leave blank if unknown."),
        ("event_type",     "See EVENT TYPE VOCABULARY below. Pick the most specific type."),
        ("procedure_ref",  "Procedure ID number (1–15) from the PROCEDURES table below. Leave blank if unclear."),
        ("description",    "In French. 1–2 sentences describing the specific event."),
        ("amount_eur",     "EUR amount as a number (e.g. 3500.00). Only for billing events. Leave blank otherwise."),
        ("significance",   "high | medium | low — how important is this event to the case?"),
    ]
    for col_name, desc in events_cols:
        w(r, 1, col_name, bold=True)
        w(r, 2, desc, wrap=True)
        ws.row_dimensions[r].height = 22
        r += 1

    # ── Analysis sheet columns ────────────────────────────────────────────────
    r += 1
    w(r, 1, "ANALYSIS SHEET — COLUMNS", bold=True, color=_ORANGE)
    r += 1
    analysis_cols = [
        ("email_id",           "Pre-filled. Do not change."),
        ("mood_valence",       f"SENT emails only. Leave blank for received. Values: {_LEGAL_MOOD_VALUES}"),
        ("mood_intensity",     "SENT emails only. 1 (routine) → 5 (peak emotion/crisis). Leave blank for received."),
        ("urgency",            f"SENT emails only. Values: {_LEGAL_URGENCY_VALUES}. Leave blank for received."),
        ("key_concern",        "SENT emails only. In French: what is Gaël's main concern in this email? Leave blank for received."),
        ("trust_in_lawyer",    f"SENT emails only. Values: {_LEGAL_TRUST_VALUES}. 'questioning' = Gaël pushes back or doubts. Leave blank for received."),
        ("father_role_stress", "SENT only. none | mild | significant — is Gaël expressing stress about being a father (children, custody, parental alienation, distance)? Leave blank for received."),
        ("financial_stress",   "SENT only. none | mild | significant — is Gaël expressing financial stress (cost, affordability, billing)? Leave blank for received."),
        ("lawyer_stance",      f"RECEIVED emails only. Values: {_LEGAL_STANCE_VALUES}. Leave blank for sent."),
        ("strategy_signal",    "RECEIVED only. In French: what strategic direction is the lawyer signaling (1 sentence)? Leave blank for sent."),
        ("action_required",    "RECEIVED only. In French: what action is the lawyer asking Gaël to do? Leave blank if none. Leave blank for sent."),
        ("risk_signal",        f"RECEIVED only. Values: {_LEGAL_RISK_VALUES}. Leave blank for sent."),
        ("procedure_ref",      "Procedure ID number (1–15). Use the most relevant procedure for this email."),
        ("persons_mentioned",  "Comma-separated names of judges, experts, opposing counsel, social workers, or other key persons named in this email. Leave blank if none."),
        ("amounts_mentioned",  "Comma-separated amounts with labels, e.g. '3500€ provision, 850€ frais huissier'. Leave blank if none."),
        ("children_mentioned", f"Values: {_LEGAL_CHILDREN_VALUES}. Which child(ren) are mentioned?"),
    ]
    for col_name, desc in analysis_cols:
        w(r, 1, col_name, bold=True)
        w(r, 2, desc, wrap=True)
        ws.row_dimensions[r].height = 22
        r += 1

    # ── Event type vocabulary ─────────────────────────────────────────────────
    r += 1
    w(r, 1, "EVENT TYPE VOCABULARY", bold=True, color=_NAVY)
    r += 1
    vocab_groups = [
        ("Billing",        ["invoice_issued", "payment_request", "payment_confirmed",
                            "fee_estimate", "expense_note", "cost_warning", "retainer_requested"]),
        ("Filings",        ["conclusions_filed", "requete_filed", "assignation", "appeal_filed",
                            "document_communicated", "constitution_avocat", "desistement",
                            "signification", "consignation", "incident_filed"]),
        ("Court",          ["hearing_scheduled", "hearing_occurred", "hearing_postponed",
                            "hearing_cancelled", "judgment_rendered", "ordonnance_rendered",
                            "arret_rendered", "expert_appointed", "expert_report_delivered",
                            "deadline_set", "mise_en_etat"]),
        ("Strategy",       ["strategy_decision", "evidence_discussed", "settlement_offer_received",
                            "settlement_offer_made", "settlement_rejected", "adverse_move",
                            "judge_observation", "case_assessment", "client_instruction",
                            "lawyer_recommendation"]),
        ("Admin",          ["meeting_scheduled", "meeting_occurred", "document_exchange",
                            "procuration", "change_of_lawyer"]),
    ]
    for group_name, types in vocab_groups:
        w(r, 1, group_name, bold=True, color=_GRAY, italic=True)
        w(r, 2, "  |  ".join(types))
        r += 1

    # ── Procedures reference ──────────────────────────────────────────────────
    r += 1
    w(r, 1, "PROCEDURES (use ID in procedure_ref)", bold=True, color=_NAVY)
    r += 1
    for proc in procedures:
        rg = proc["case_number"] or "—"
        w(r, 1, f"#{proc['id']}", bold=True)
        w(r, 2, f"{proc['name']}  ({rg})")
        r += 1

    # ── Lawyers reference ─────────────────────────────────────────────────────
    r += 1
    w(r, 1, "LAWYERS", bold=True, color=_NAVY)
    r += 1
    role_label = {"my_lawyer": "Gaël's lawyer", "her_lawyer": "Maud's lawyer",
                  "opposing_counsel": "opposing counsel", "notaire": "notaire"}
    for lw in lawyers:
        w(r, 1, role_label.get(lw["role"], lw["role"]), bold=True, color=_GRAY, italic=True)
        w(r, 2, lw["name"])
        r += 1

    # ── Import command ────────────────────────────────────────────────────────
    r += 1
    w(r, 1, "IMPORT COMMAND", bold=True, color=_NAVY)
    w(r, 2,
      "python cli.py analyze import-results <this_file.xlsx> "
      "--type legal_analysis --provider openai --model gpt-4o")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 100


def _build_legal_emails_sheet(wb, email_data):
    ws = wb.create_sheet("Emails")
    headers = ["email_id", "date", "direction", "contact_name", "contact_role", "subject", "content"]

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = PatternFill("solid", fgColor=_BLUE_HDR)
        cell.font = Font(color=_WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    for row_idx, em in enumerate(email_data, start=2):
        ws.cell(row=row_idx, column=1, value=em["id"])
        ws.cell(row=row_idx, column=2, value=em["date"])
        ws.cell(row=row_idx, column=3, value=em["direction"])
        ws.cell(row=row_idx, column=4, value=em["contact_name"])
        ws.cell(row=row_idx, column=5, value=em["contact_role"])
        ws.cell(row=row_idx, column=6, value=em["subject"])
        cell = ws.cell(row=row_idx, column=7, value=em["content"])
        cell.alignment = Alignment(wrap_text=False, vertical="top")

    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 11
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 38
    ws.column_dimensions["G"].width = 70
    ws.freeze_panes = "A2"


def _build_legal_events_sheet(wb, email_data):
    """Pre-populate one row per email with email_id (blue). Output cells are yellow."""
    ws = wb.create_sheet("Events")
    headers = ["email_id", "event_date", "event_type", "procedure_ref",
               "description", "amount_eur", "significance"]

    blue_fill   = PatternFill("solid", fgColor=_BLUE_HDR)
    orange_fill = PatternFill("solid", fgColor=_ORANGE)
    yellow_fill = PatternFill("solid", fgColor=_YELLOW_BG)

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = orange_fill
        cell.font = Font(color=_WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    for row_idx, em in enumerate(email_data, start=2):
        # email_id: blue (pre-filled, read-only)
        id_cell = ws.cell(row=row_idx, column=1, value=em["id"])
        id_cell.fill = blue_fill
        id_cell.font = Font(color=_WHITE, bold=True)
        # output cells: yellow blank
        for col_idx in range(2, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx, value="").fill = yellow_fill

    ws.column_dimensions["A"].width = 10   # email_id
    ws.column_dimensions["B"].width = 13   # event_date
    ws.column_dimensions["C"].width = 26   # event_type
    ws.column_dimensions["D"].width = 14   # procedure_ref
    ws.column_dimensions["E"].width = 70   # description
    ws.column_dimensions["F"].width = 13   # amount_eur
    ws.column_dimensions["G"].width = 13   # significance
    ws.freeze_panes = "B2"


def _build_legal_analysis_sheet(wb, email_data):
    """Pre-populate one row per email with email_id (blue). All other cells are yellow."""
    ws = wb.create_sheet("Analysis")
    headers = [
        "email_id",
        "mood_valence", "mood_intensity", "urgency", "key_concern",
        "trust_in_lawyer", "father_role_stress", "financial_stress",
        "lawyer_stance", "strategy_signal", "action_required", "risk_signal",
        "procedure_ref", "persons_mentioned", "amounts_mentioned", "children_mentioned",
    ]

    blue_fill   = PatternFill("solid", fgColor=_BLUE_HDR)
    orange_fill = PatternFill("solid", fgColor=_ORANGE)
    yellow_fill = PatternFill("solid", fgColor=_YELLOW_BG)

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = orange_fill
        cell.font = Font(color=_WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    for row_idx, em in enumerate(email_data, start=2):
        id_cell = ws.cell(row=row_idx, column=1, value=em["id"])
        id_cell.fill = blue_fill
        id_cell.font = Font(color=_WHITE, bold=True)
        for col_idx in range(2, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx, value="").fill = yellow_fill

    # Column widths
    widths = {
        "A": 10, "B": 16, "C": 16, "D": 13, "E": 50,
        "F": 16, "G": 20, "H": 18,
        "I": 16, "J": 50, "K": 40, "L": 13,
        "M": 14, "N": 40, "O": 35, "P": 20,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "B2"


def _build_legal_meta_sheet(wb, email_count):
    ws = wb.create_sheet("_meta")
    ws["A1"], ws["B1"] = "analysis_type", "legal_analysis"
    ws["A2"], ws["B2"] = "export_date",   datetime.now().isoformat()
    ws["A3"], ws["B3"] = "email_count",   email_count
    ws["A4"], ws["B4"] = "corpus",        "legal"
    ws["A5"], ws["B5"] = "schema_version", "1"


# ── Contradictions export (topic-based, special format) ───────────────────────

_CONTRADICTION_PATTERNS = [
    "gaslighting", "emotional_weaponization", "financial_coercion", "legal_threats",
    "children_instrumentalization", "guilt_tripping", "projection", "false_victimhood",
    "moving_goalposts", "silent_treatment_threat",
]

_VALID_TOPICS = [
    "enfants", "finances", "école", "logement", "vacances", "santé",
    "procédure", "éducation", "activités", "divorce", "contradictions", "famille",
]


def export_contradictions_batch(
    conn: sqlite3.Connection,
    output_path: Path,
    topic: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = 600,
) -> tuple:
    """
    Export email summaries grouped by topic for contradiction detection.

    Uses classification summaries (not full delta_text) — much more token-efficient.
    Output sheet has pre-populated headers for ChatGPT to fill contradiction pairs.

    Returns (output_path, email_count).
    """
    if not _HAS_OPENPYXL:
        raise ImportError("openpyxl required: pip install openpyxl")

    # Build filter
    wheres = [
        "ar.result_json IS NOT NULL",
        "ru.analysis_type = 'classify'",
        "ru.status IN ('complete', 'partial')",
        "t.name NOT IN ('trop_court', 'non_classifiable')",
    ]
    params: list = []

    if topic:
        wheres.append("t.name = ?")
        params.append(topic)
    if date_from:
        wheres.append("e.date >= ?")
        params.append(date_from)
    if date_to:
        wheres.append("e.date <= ?")
        params.append(date_to)

    rows = conn.execute(f"""
        SELECT DISTINCT e.id, e.date, e.direction, e.subject,
               JSON_EXTRACT(ar.result_json, '$.summary') AS summary,
               c.name AS contact_name,
               GROUP_CONCAT(DISTINCT t.name) AS topics
        FROM emails e
        JOIN email_topics et ON et.email_id = e.id
        JOIN topics t ON t.id = et.topic_id
        JOIN analysis_results ar ON ar.email_id = e.id
        JOIN analysis_runs ru ON ru.id = ar.run_id
        LEFT JOIN contacts c ON c.id = e.contact_id
        WHERE {' AND '.join(wheres)}
        GROUP BY e.id
        ORDER BY e.date ASC
        LIMIT {limit}
    """, params).fetchall()

    if not rows:
        raise ValueError("No classified emails found for the given filters.")

    wb = openpyxl.Workbook()

    # ── Instructions sheet ────────────────────────────────────────────────────
    ws_inst = wb.active
    ws_inst.title = "Instructions"

    ws_inst["A1"] = "MyYahooEmails — Contradiction Detection Batch"
    ws_inst["A1"].font = Font(bold=True, size=15, color=_NAVY)
    ws_inst["A2"] = (
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}   |   "
        f"{len(rows)} emails   |   topic: {topic or 'all'}   |   "
        f"period: {date_from or 'start'} → {date_to or 'end'}"
    )
    ws_inst["A2"].font = Font(size=11, color=_GRAY)

    ws_inst["A4"] = "TASK"
    ws_inst["A4"].font = Font(bold=True, color=_NAVY)
    ws_inst["B4"] = (
        "You are a forensic legal analyst reviewing French divorce emails. "
        "Read ALL email summaries in the 'Emails' sheet (sorted by date). "
        "Identify pairs of emails where the content directly contradicts each other — "
        "same person making incompatible claims, broken commitments, conflicting facts or dates. "
        "Fill the 'Contradictions' output sheet: one row per contradiction found. "
        "Do NOT modify the Emails sheet."
    )
    ws_inst["B4"].alignment = Alignment(wrap_text=True, vertical="top")
    ws_inst["B4"].font = Font(size=11)
    ws_inst.row_dimensions[4].height = 60

    ws_inst["A6"] = "CONTRADICTION TYPES"
    ws_inst["A6"].font = Font(bold=True, color=_NAVY)
    ws_inst["B6"] = (
        "intra-sender: both emails from the same person, contradicting themselves\n"
        "cross-sender: one email per party, one directly refutes the other's claim"
    )
    ws_inst["B6"].alignment = Alignment(wrap_text=True)
    ws_inst.row_dimensions[6].height = 35

    ws_inst["A8"] = "SEVERITY"
    ws_inst["A8"].font = Font(bold=True, color=_NAVY)
    ws_inst["B8"] = (
        "high: clear lie, perjury risk, or directly falsifiable factual claim\n"
        "medium: significant discrepancy in facts, dates, or commitments\n"
        "low: minor inconsistency or change of position over time"
    )
    ws_inst["B8"].alignment = Alignment(wrap_text=True)
    ws_inst.row_dimensions[8].height = 50

    ws_inst["A10"] = "IMPORT COMMAND"
    ws_inst["A10"].font = Font(bold=True, color=_NAVY)
    ws_inst["B10"] = (
        "python cli.py analyze import-results <file.xlsx> --type contradictions "
        "--provider openai --model gpt-5.4-thinking"
    )

    ws_inst.column_dimensions["A"].width = 22
    ws_inst.column_dimensions["B"].width = 90

    # ── Emails sheet (read-only summaries) ────────────────────────────────────
    ws_emails = wb.create_sheet("Emails")
    email_headers = ["email_id", "date", "direction", "contact", "subject", "summary", "topics"]
    for col_idx, h in enumerate(email_headers, start=1):
        cell = ws_emails.cell(row=1, column=col_idx, value=h)
        cell.fill = PatternFill("solid", fgColor=_BLUE_HDR)
        cell.font = Font(color=_WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center")
    ws_emails.row_dimensions[1].height = 22

    for row_idx, row in enumerate(rows, start=2):
        ws_emails.cell(row=row_idx, column=1, value=row["id"])
        ws_emails.cell(row=row_idx, column=2, value=str(row["date"])[:10] if row["date"] else "")
        ws_emails.cell(row=row_idx, column=3, value=row["direction"] or "")
        ws_emails.cell(row=row_idx, column=4, value=row["contact_name"] or "")
        ws_emails.cell(row=row_idx, column=5, value=row["subject"] or "")
        ws_emails.cell(row=row_idx, column=6, value=row["summary"] or "")
        ws_emails.cell(row=row_idx, column=7, value=row["topics"] or "")

    ws_emails.column_dimensions["A"].width = 10
    ws_emails.column_dimensions["B"].width = 12
    ws_emails.column_dimensions["C"].width = 11
    ws_emails.column_dimensions["D"].width = 16
    ws_emails.column_dimensions["E"].width = 38
    ws_emails.column_dimensions["F"].width = 65
    ws_emails.column_dimensions["G"].width = 30
    ws_emails.freeze_panes = "A2"

    # ── Contradictions output sheet ───────────────────────────────────────────
    ws_contra = wb.create_sheet("Contradictions")
    contra_headers = ["email_id_a", "email_id_b", "scope", "topic", "severity", "explanation"]
    header_notes = [
        "ID of first email",
        "ID of second email",
        "intra-sender OR cross-sender",
        "Main topic (e.g. enfants, finances)",
        "high | medium | low",
        "Explanation in French (max 3 sentences, quote key phrases)",
    ]
    for col_idx, (h, note) in enumerate(zip(contra_headers, header_notes), start=1):
        cell = ws_contra.cell(row=1, column=col_idx, value=h)
        cell.fill = PatternFill("solid", fgColor=_ORANGE)
        cell.font = Font(color=_WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.comment = None  # openpyxl doesn't support comments easily, use row 2 for hints
    ws_contra.row_dimensions[1].height = 22

    # Row 2: hint row
    yellow = PatternFill("solid", fgColor=_YELLOW_BG)
    for col_idx, note in enumerate(header_notes, start=1):
        cell = ws_contra.cell(row=2, column=col_idx, value=f"← {note}")
        cell.fill = yellow
        cell.font = Font(italic=True, color=_GRAY, size=9)

    # Pre-fill 50 empty yellow rows
    for row_idx in range(3, 53):
        for col_idx in range(1, 7):
            ws_contra.cell(row=row_idx, column=col_idx, value="").fill = yellow

    ws_contra.column_dimensions["A"].width = 12
    ws_contra.column_dimensions["B"].width = 12
    ws_contra.column_dimensions["C"].width = 16
    ws_contra.column_dimensions["D"].width = 18
    ws_contra.column_dimensions["E"].width = 10
    ws_contra.column_dimensions["F"].width = 80
    ws_contra.freeze_panes = "A3"

    # ── Meta sheet ────────────────────────────────────────────────────────────
    ws_meta = wb.create_sheet("_meta")
    ws_meta["A1"], ws_meta["B1"] = "analysis_type", "contradictions"
    ws_meta["A2"], ws_meta["B2"] = "export_date", datetime.now().isoformat()
    ws_meta["A3"], ws_meta["B3"] = "email_count", len(rows)
    ws_meta["A4"], ws_meta["B4"] = "topic", topic or "all"
    ws_meta["A5"], ws_meta["B5"] = "date_from", date_from or ""
    ws_meta["A6"], ws_meta["B6"] = "date_to", date_to or ""

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path, len(rows)
